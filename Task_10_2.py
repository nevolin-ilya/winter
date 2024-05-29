import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\NevolinIS\Desktop\Python\Домашка\Task_10_2.xlsx")
print(wb.sheetnames)
ws = wb.active
su = 0
lst = []
for i in range(1,4):
    su += ws.cell(i, 2).value
    lst.append((ws.cell(i, 1).value, ws.cell(i, 2).value,))

s = sorted(lst, key= lambda x: -x[1])
print(s)
print(su)

wb.create_sheet("Newsheets")
ws = wb["Newsheets"]
for row_idx, (col1, col2) in enumerate(s, start=1):
    print(ws.cell(row=row_idx, column=1, value=col1))
    print(ws.cell(row=row_idx, column=2, value=col2))
ws.cell(4, 2).value = su
print(ws.cell(4, 2))
wb.save(r"C:\Users\NevolinIS\Desktop\Python\Домашка\Task_10_2.xlsx")
wb.close()