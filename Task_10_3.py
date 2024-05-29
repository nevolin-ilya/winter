import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\NevolinIS\Desktop\Python\Домашка\Task_10_3.xlsx")
print(wb.sheetnames)
ws = wb.active

su = 0
lst = []
for i in range(1,5):
    su += ws.cell(i, 2).value
    lst.append(ws.cell(i, 2).value)

maximum = max(lst)
minimum = min(lst)
print(maximum)
print(minimum)
print(su)
sred = int(su/4)
print(sred)
med = lst[1] + (lst[2] - lst[1])/2
print(int(med))

wb.create_sheet("Newsheets")
ws = wb["Newsheets"]
print(ws.cell(row=1, column=1, value="Максимальное значение"), ws.cell(row=1, column=2, value=maximum))
print(ws.cell(row=2, column=1, value="Минимальное значение"), ws.cell(row=2, column=2, value=minimum))
print(ws.cell(row=3, column=1, value="Сумма"), ws.cell(row=3, column=2, value=su))
print(ws.cell(row=4, column=1, value="Среднее арифмитическое"), ws.cell(row=4, column=2, value=sred))
print(ws.cell(row=5, column=1, value="Медиана"), ws.cell(row=5, column=2, value=med))

wb.save(r"C:\Users\NevolinIS\Desktop\Python\Домашка\Task_10_3.xlsx")
wb.close()